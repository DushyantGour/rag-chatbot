import os
import re
from pathlib import Path

# Document loaders
import pypdf
import docx
import openpyxl

# Embeddings + Vector DB
from sentence_transformers import SentenceTransformer
import chromadb

# Groq
from groq import Groq
from dotenv import load_dotenv

load_dotenv()

# ── Config ──────────────────────────────────────────
DOCS_FOLDER = "documents"
CHROMA_FOLDER = "chroma_db"
EMBED_MODEL = "all-MiniLM-L6-v2"       # small, fast, local
GROQ_MODEL = "llama-3.1-8b-instant"           # fast & free on Groq
CHUNK_SIZE = 500                         # characters per chunk
CHUNK_OVERLAP = 50

# ── Init ─────────────────────────────────────────────
embedder = SentenceTransformer(EMBED_MODEL)
chroma_client = chromadb.PersistentClient(path=CHROMA_FOLDER)
collection = chroma_client.get_or_create_collection(name="documents")
groq_client = Groq(api_key=os.getenv("GROQ_API_KEY"))


# ── 1. File Readers ──────────────────────────────────

def read_pdf(path):
    text = ""
    reader = pypdf.PdfReader(path)
    for page in reader.pages:
        text += page.extract_text() or ""
    return text

def read_docx(path):
    doc = docx.Document(path)
    return "\n".join([p.text for p in doc.paragraphs])

def read_xlsx(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    text = ""
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        text += f"\n[Sheet: {sheet}]\n"
        for row in ws.iter_rows(values_only=True):
            row_text = " | ".join([str(c) for c in row if c is not None])
            if row_text.strip():
                text += row_text + "\n"
    return text

def load_documents():
    docs = []
    for file in Path(DOCS_FOLDER).iterdir():
        try:
            if file.suffix.lower() == ".pdf":
                text = read_pdf(file)
            elif file.suffix.lower() == ".docx":
                text = read_docx(file)
            elif file.suffix.lower() in [".xlsx", ".xls"]:
                text = read_xlsx(file)
            else:
                continue
            docs.append({"filename": file.name, "text": text})
            print(f"✅ Loaded: {file.name}")
        except Exception as e:
            print(f"❌ Error loading {file.name}: {e}")
    return docs


# ── 2. Chunking ───────────────────────────────────────

def chunk_text(text, filename):
    chunks = []
    start = 0
    idx = 0
    while start < len(text):
        end = start + CHUNK_SIZE
        chunk = text[start:end]
        if chunk.strip():
            chunks.append({
                "id": f"{filename}_{idx}",
                "text": chunk,
                "source": filename
            })
            idx += 1
        start = end - CHUNK_OVERLAP
    return chunks


# ── 3. Embed & Store ──────────────────────────────────

def index_documents():
    # Skip if already indexed
    if collection.count() > 0:
        print(f"📦 Vector DB already has {collection.count()} chunks. Skipping indexing.")
        return

    docs = load_documents()
    if not docs:
        print("⚠️  No documents found in /documents folder!")
        return

    all_chunks = []
    for doc in docs:
        all_chunks.extend(chunk_text(doc["text"], doc["filename"]))

    texts = [c["text"] for c in all_chunks]
    ids = [c["id"] for c in all_chunks]
    metadatas = [{"source": c["source"]} for c in all_chunks]

    print(f"🔍 Embedding {len(texts)} chunks...")
    embeddings = embedder.encode(texts, show_progress_bar=True).tolist()

    collection.add(documents=texts, embeddings=embeddings, ids=ids, metadatas=metadatas)
    print(f"✅ Indexed {len(texts)} chunks into ChromaDB.")


# ── 4. Query ──────────────────────────────────────────

def query(user_question, top_k=5):
    # Embed the question
    q_embedding = embedder.encode([user_question]).tolist()

    # Search ChromaDB
    results = collection.query(query_embeddings=q_embedding, n_results=top_k)
    chunks = results["documents"][0]
    sources = [m["source"] for m in results["metadatas"][0]]

    # Build context
    context = "\n\n".join([f"[{src}]\n{chunk}" for src, chunk in zip(sources, chunks)])

    # Prompt
    prompt = f"""You are a helpful assistant. Answer the user's question using ONLY the context below.
If the answer is not in the context, say "I couldn't find that in the documents."

Context:
{context}

Question: {user_question}
Answer:"""

    # Call Groq
    response = groq_client.chat.completions.create(
        model=GROQ_MODEL,
        messages=[{"role": "user", "content": prompt}],
        temperature=0.2,
    )

    answer = response.choices[0].message.content
    unique_sources = list(set(sources))
    return answer, unique_sources


# ── 5. Re-index (force refresh) ───────────────────────

def reindex():
    collection.delete(ids=collection.get()["ids"])
    print("🗑️  Cleared old index.")
    index_documents()